import netmiko
import pprint
from getpass import getpass
import textfsm
import openpyxl

ip = input("Insert IP: ")
username = input("Insert username: ")
password = getpass("Insert password: ")

router = {
	'ip' : ip,
	'username' : username,
	'password' : password,
	'device_type': 'cisco_ios'
}

net_connect = netmiko.ConnectHandler(**router)

string_inventory = net_connect.send_command('show inventory')

template_file = open("inventory.template")
template = textfsm.TextFSM(template_file)
result_template = template.ParseText(string_inventory)

print(result_template)
row =2	
wb = openpyxl.Workbook()
ws = wb.active

ws.cell(row=1, column=1, value="Name")
ws.cell(row=1, column=2, value="Description")
ws.cell(row=1, column=3, value="PID")
ws.cell(row=1, column=4, value="VID")
ws.cell(row=1, column=5, value="Serial Number")

for x in result_template:
	ws.cell(row=row, column=1, value =x[0])
	ws.cell(row=row, column=2, value =x[1])
	ws.cell(row=row, column=3, value =x[2])
	ws.cell(row=row, column=4, value =x[3])
	ws.cell(row=row, column=5, value =x[4])
	row += 1
wb.save("sample.xlsx")